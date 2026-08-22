import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def get_floor_name(team_id):
    if team_id.startswith("NEX0"):
        return "Ground Floor"
    elif team_id.startswith("NEX1"):
        return "First Floor"
    elif team_id.startswith("NEX2"):
        return "Second Floor"
    elif team_id.startswith("NEX3"):
        return "Online / Virtual"
    return "Main Venue"

def generate_excel(db_path, output_excel_path):
    with open(db_path, "r") as f:
        db = json.load(f)

    teams = db.get("teams", [])
    
    cyber_teams = sorted([t for t in teams if t.get("track") == "Cyber Security"], key=lambda x: x["id"])
    med_teams = sorted([t for t in teams if t.get("track") == "Med-Tech"], key=lambda x: x["id"])
    all_teams_sorted = sorted(teams, key=lambda x: x["id"])

    wb = Workbook()

    # Reusable styles
    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Calibri", size=10, italic=True, color="E2E8F0")
    font_section = Font(name="Calibri", size=12, bold=True, color="0F172A")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True, color="0F172A")
    font_regular = Font(name="Calibri", size=10, color="1E293B")
    
    font_cyber = Font(name="Calibri", size=10, bold=True, color="6B21A8")
    font_med = Font(name="Calibri", size=10, bold=True, color="047857")

    fill_title = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    fill_header_main = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_header_cyber = PatternFill(start_color="581C87", end_color="581C87", fill_type="solid")
    fill_header_med = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_total = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    thick_bottom = Border(bottom=Side(style='medium', color='0F172A'))

    # ----------------------------------------------------
    # TAB 1: Overview & Summary
    # ----------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = True

    # Title Block
    ws_summary.merge_cells("A1:F1")
    ws_summary["A1"] = "NEXORA 2026 HACKATHON — ADMIN REPORT & TRACK SPLIT"
    ws_summary["A1"].font = font_title
    ws_summary["A1"].fill = fill_title
    ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 35

    ws_summary.merge_cells("A2:F2")
    ws_summary["A2"] = "Official Event Statistics & Venue Distribution Overview"
    ws_summary["A2"].font = font_subtitle
    ws_summary["A2"].fill = fill_title
    ws_summary["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[2].height = 20

    # KPI Block
    ws_summary["A4"] = "Key Performance Indicators"
    ws_summary["A4"].font = font_section

    kpi_headers = ["Metric", "Value", "Notes"]
    for col_idx, h in enumerate(kpi_headers, 1):
        cell = ws_summary.cell(row=5, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header_main
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    kpis = [
        ("Total Registered Teams", len(teams), "100% Fully Configured"),
        ("Estimated Participants", len(teams) * 4, "Based on 4 Members / Team"),
        ("Cyber Security Track Teams", len(cyber_teams), f"{round(len(cyber_teams)/len(teams)*100, 1)}% of total"),
        ("Med-Tech Track Teams", len(med_teams), f"{round(len(med_teams)/len(teams)*100, 1)}% of total"),
        ("On-Site Teams (Floors 0-2)", len(teams) - 23, "Ground, 1st & 2nd Floors"),
        ("Online Virtual Teams", 23, "NEX3001 - NEX3023"),
    ]

    for idx, (m, v, n) in enumerate(kpis, start=6):
        c1 = ws_summary.cell(row=idx, column=1, value=m)
        c2 = ws_summary.cell(row=idx, column=2, value=v)
        c3 = ws_summary.cell(row=idx, column=3, value=n)

        c1.font = font_bold
        c2.font = font_bold
        c3.font = font_regular

        c1.border = thin_border
        c2.border = thin_border
        c3.border = thin_border

        c2.alignment = Alignment(horizontal="center")

    # Venue Breakdown Table
    ws_summary["A14"] = "Venue & Floor-wise Track Breakdown"
    ws_summary["A14"].font = font_section

    venue_headers = ["Venue / Location", "ID Range", "Cyber Security", "Med-Tech", "Total Teams", "Track Split %"]
    for col_idx, h in enumerate(venue_headers, 1):
        cell = ws_summary.cell(row=15, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header_main
        cell.alignment = Alignment(horizontal="center", vertical="center")

    floors_data = [
        ("Ground Floor", "NEX0001 - NEX0046", 23, 23, 46, "50% / 50%"),
        ("First Floor", "NEX1001 - NEX1047", 24, 23, 47, "51% / 49%"),
        ("Second Floor", "NEX2001 - NEX2058", 29, 29, 58, "50% / 50%"),
        ("Online / Virtual", "NEX3001 - NEX3023", 11, 12, 23, "48% / 52%"),
    ]

    for r_idx, f_row in enumerate(floors_data, start=16):
        for c_idx, val in enumerate(f_row, start=1):
            cell = ws_summary.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_bold if c_idx == 1 else font_regular
            cell.border = thin_border
            if c_idx in [3, 4, 5, 6]:
                cell.alignment = Alignment(horizontal="center")

    # Total Row
    tot_row = 20
    ws_summary.cell(row=tot_row, column=1, value="GRAND TOTAL").font = font_bold
    ws_summary.cell(row=tot_row, column=2, value="ALL VENUES").font = font_bold
    ws_summary.cell(row=tot_row, column=3, value=len(cyber_teams)).font = font_cyber
    ws_summary.cell(row=tot_row, column=4, value=len(med_teams)).font = font_med
    ws_summary.cell(row=tot_row, column=5, value=len(teams)).font = font_bold
    ws_summary.cell(row=tot_row, column=6, value="50% / 50%").font = font_bold

    for col_idx in range(1, 7):
        cell = ws_summary.cell(row=tot_row, column=col_idx)
        cell.fill = fill_total
        cell.border = thin_border

    # ----------------------------------------------------
    # TAB 2: Cyber Security Track
    # ----------------------------------------------------
    ws_cyber = wb.create_sheet(title="Cyber Security Track")
    ws_cyber.views.sheetView[0].showGridLines = True

    cyber_cols = ["S.No", "Team ID", "Team Name", "Team Leader", "Phone Number", "Venue / Floor", "Problem Statement PDF", "Status"]
    for col_idx, h in enumerate(cyber_cols, 1):
        cell = ws_cyber.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header_cyber
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_cyber.row_dimensions[1].height = 25

    for idx, t in enumerate(cyber_teams, start=1):
        r = idx + 1
        ws_cyber.cell(row=r, column=1, value=idx).alignment = Alignment(horizontal="center")
        ws_cyber.cell(row=r, column=2, value=t["id"]).font = font_cyber
        ws_cyber.cell(row=r, column=3, value=t["name"]).font = font_bold
        ws_cyber.cell(row=r, column=4, value=t.get("leaderName", "N/A"))
        ws_cyber.cell(row=r, column=5, value=t.get("leaderPhone", "N/A")).alignment = Alignment(horizontal="center")
        ws_cyber.cell(row=r, column=6, value=get_floor_name(t["id"]))
        ws_cyber.cell(row=r, column=7, value="Cyber Security Official PS (2.pdf)")
        ws_cyber.cell(row=r, column=8, value=t.get("status", "In Progress")).alignment = Alignment(horizontal="center")

        fill_row = fill_zebra if r % 2 == 0 else PatternFill(fill_type=None)
        for col_idx in range(1, 9):
            cell = ws_cyber.cell(row=r, column=col_idx)
            cell.border = thin_border
            if fill_row.fill_type:
                cell.fill = fill_row

    # ----------------------------------------------------
    # TAB 3: Med-Tech Track
    # ----------------------------------------------------
    ws_med = wb.create_sheet(title="Med-Tech Track")
    ws_med.views.sheetView[0].showGridLines = True

    med_cols = ["S.No", "Team ID", "Team Name", "Team Leader", "Phone Number", "Venue / Floor", "Problem Statement PDF", "Status"]
    for col_idx, h in enumerate(med_cols, 1):
        cell = ws_med.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header_med
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_med.row_dimensions[1].height = 25

    for idx, t in enumerate(med_teams, start=1):
        r = idx + 1
        ws_med.cell(row=r, column=1, value=idx).alignment = Alignment(horizontal="center")
        ws_med.cell(row=r, column=2, value=t["id"]).font = font_med
        ws_med.cell(row=r, column=3, value=t["name"]).font = font_bold
        ws_med.cell(row=r, column=4, value=t.get("leaderName", "N/A"))
        ws_med.cell(row=r, column=5, value=t.get("leaderPhone", "N/A")).alignment = Alignment(horizontal="center")
        ws_med.cell(row=r, column=6, value=get_floor_name(t["id"]))
        ws_med.cell(row=r, column=7, value="Med-Tech Official PS (1.pdf)")
        ws_med.cell(row=r, column=8, value=t.get("status", "In Progress")).alignment = Alignment(horizontal="center")

        fill_row = fill_zebra if r % 2 == 0 else PatternFill(fill_type=None)
        for col_idx in range(1, 9):
            cell = ws_med.cell(row=r, column=col_idx)
            cell.border = thin_border
            if fill_row.fill_type:
                cell.fill = fill_row

    # ----------------------------------------------------
    # TAB 4: All Teams Master Roster
    # ----------------------------------------------------
    ws_all = wb.create_sheet(title="All Teams Master Roster")
    ws_all.views.sheetView[0].showGridLines = True

    all_cols = ["S.No", "Team ID", "Team Name", "Assigned Track", "Team Leader", "Phone Number", "Venue / Floor", "Roster Status"]
    for col_idx, h in enumerate(all_cols, 1):
        cell = ws_all.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header_main
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_all.row_dimensions[1].height = 25

    for idx, t in enumerate(all_teams_sorted, start=1):
        r = idx + 1
        is_cyber = t.get("track") == "Cyber Security"
        ws_all.cell(row=r, column=1, value=idx).alignment = Alignment(horizontal="center")
        ws_all.cell(row=r, column=2, value=t["id"]).font = font_cyber if is_cyber else font_med
        ws_all.cell(row=r, column=3, value=t["name"]).font = font_bold
        ws_all.cell(row=r, column=4, value=t.get("track")).font = font_cyber if is_cyber else font_med
        ws_all.cell(row=r, column=5, value=t.get("leaderName", "N/A"))
        ws_all.cell(row=r, column=6, value=t.get("leaderPhone", "N/A")).alignment = Alignment(horizontal="center")
        ws_all.cell(row=r, column=7, value=get_floor_name(t["id"]))
        ws_all.cell(row=r, column=8, value="Locked" if t.get("isRosterLocked") else "Editable").alignment = Alignment(horizontal="center")

        fill_row = fill_zebra if r % 2 == 0 else PatternFill(fill_type=None)
        for col_idx in range(1, 9):
            cell = ws_all.cell(row=r, column=col_idx)
            cell.border = thin_border
            if fill_row.fill_type:
                cell.fill = fill_row

    # Auto-adjust Column Widths across all worksheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.coordinate in ["A1", "A2", "A14"]:  # Skip title merges
                    continue
                max_len = max(max_len, len(val_str))
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(output_excel_path)
    print(f"Excel report successfully generated at: {output_excel_path}")

if __name__ == "__main__":
    db_file = "./data/final_db.json"
    out_excel = "./public/uploads/Nexora_Admin_Track_Report.xlsx"
    os.makedirs(os.path.dirname(out_excel), exist_ok=True)
    generate_excel(db_file, out_excel)
    
    # Also save to root directory for easy local access
    root_excel = "./Nexora_Admin_Track_Report.xlsx"
    generate_excel(db_file, root_excel)
